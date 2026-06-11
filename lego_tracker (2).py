"""
LEGO Market Tracker
Pobiera publiczne dane o ofertach z Allegro.
Wymaga aplikacji typu: "Aplikacja ma dostep do przegladarki" (authorization_code)
z uzyciem Client Credentials flow.
"""

import os, sys, json, base64, urllib.request, urllib.parse
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

EXCEL_FILE = "LEGO_Monitor.xlsx"

C_LIGHT  = "D6EAF8"
C_WHITE  = "FFFFFF"
C_BORDER = "BDC3C7"

def _border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _style_row(ws, row, ncols, alt):
    bg = C_LIGHT if alt else C_WHITE
    for col in range(1, ncols + 1):
        c = ws.cell(row, col)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(name="Arial", size=10)
        c.border    = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")

# ── Token: Client Credentials (dziala tylko z typem authorization_code) ──────
def get_token() -> str:
    CLIENT_ID     = os.environ.get("ALLEGRO_CLIENT_ID", "")
    CLIENT_SECRET = os.environ.get("ALLEGRO_CLIENT_SECRET", "")

    if not CLIENT_ID or not CLIENT_SECRET:
        print("BLAD: Brak ALLEGRO_CLIENT_ID lub ALLEGRO_CLIENT_SECRET.")
        sys.exit(1)

    creds = base64.b64encode(f"{CLIENT_ID}:{CLIENT_SECRET}".encode()).decode()
    data  = urllib.parse.urlencode({"grant_type": "client_credentials"}).encode()

    req = urllib.request.Request(
        "https://allegro.pl/auth/oauth/token",
        data=data,
        headers={
            "Authorization": f"Basic {creds}",
            "Content-Type":  "application/x-www-form-urlencoded"
        }
    )
    try:
        with urllib.request.urlopen(req) as r:
            result = json.loads(r.read())
            print("Token OK.")
            return result["access_token"]
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        print(f"BLAD autoryzacji: {body}")
        sys.exit(1)

def fetch_offers(token: str, set_nr: str) -> list:
    params = urllib.parse.urlencode({
        "phrase":   f"LEGO {set_nr}",
        "limit":    100,
        "sort":     "price",
        "include":  "-all",
        "fallback": "false",
    })
    req = urllib.request.Request(
        f"https://api.allegro.pl/offers/listing?{params}",
        headers={
            "Authorization": f"Bearer {token}",
            "Accept":        "application/vnd.allegro.public.v1+json"
        }
    )
    with urllib.request.urlopen(req) as r:
        data = json.loads(r.read())
    return (data.get("items", {}).get("regular", []) +
            data.get("items", {}).get("promoted", []))

def parse_stats(items: list) -> dict:
    prices      = []
    bought_last = 0
    for it in items:
        try:
            prices.append(float(it["sellingMode"]["price"]["amount"]))
        except (KeyError, TypeError, ValueError):
            pass
        stats = it.get("stats", {})
        bought_last = max(
            bought_last,
            stats.get("boughtCount", 0) or 0,
            stats.get("popularityScore", 0) or 0,
        )
    return {
        "min_price":   round(min(prices), 2) if prices else None,
        "offer_count": len(prices),
        "bought_last": bought_last,
    }

def read_sets(wb) -> list:
    ws = wb["Lista zestawow"]
    sets = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        nr    = str(row[0]).strip() if row[0] else ""
        name  = str(row[1]).strip() if row[1] else ""
        seria = str(row[2]).strip() if row[2] else ""
        aktyw = str(row[4]).strip().upper() if len(row) > 4 and row[4] else ""
        if nr and aktyw == "TAK":
            sets.append((nr, name, seria))
    return sets

def append_row(wb, today, nr, name, seria, stats):
    ws       = wb["Dane dzienne"]
    next_row = ws.max_row + 1
    alt      = next_row % 2 == 0
    ws.cell(next_row, 1, today)
    ws.cell(next_row, 2, nr)
    ws.cell(next_row, 3, name)
    ws.cell(next_row, 4, seria)
    ws.cell(next_row, 5, stats["min_price"])
    ws.cell(next_row, 6, stats["offer_count"])
    ws.cell(next_row, 7, stats["bought_last"])
    ws.cell(next_row, 8, "OK")
    _style_row(wb["Dane dzienne"], next_row, 8, alt)
    if stats["min_price"]:
        ws.cell(next_row, 5).number_format = '#,##0.00 "zł"'

def main():
    today      = str(date.today())
    excel_path = Path(EXCEL_FILE)

    if not excel_path.exists():
        print(f"BLAD: Nie znaleziono {EXCEL_FILE}")
        sys.exit(1)

    print(f"[{today}] Pobieram token...")
    token = get_token()

    wb   = openpyxl.load_workbook(excel_path)
    sets = read_sets(wb)

    if not sets:
        print("Brak aktywnych zestawow.")
        sys.exit(0)

    print(f"Znaleziono {len(sets)} zestawow.")

    for nr, name, seria in sets:
        print(f"  -> {nr} {name}...", end=" ", flush=True)
        try:
            items = fetch_offers(token, nr)
            stats = parse_stats(items)
            append_row(wb, today, nr, name, seria, stats)
            print(f"cena: {stats['min_price']} zl | ofert: {stats['offer_count']} | kupilo: {stats['bought_last']}")
        except Exception as e:
            print(f"BLAD: {e}")
            append_row(wb, today, nr, name, seria,
                       {"min_price": None, "offer_count": 0, "bought_last": 0})

    wb.save(excel_path)
    print(f"\nZapisano: {excel_path}")

if __name__ == "__main__":
    main()
