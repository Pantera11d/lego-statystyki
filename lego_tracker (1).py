"""
LEGO Market Tracker
Pobiera dane z Allegro i dopisuje do LEGO_Monitor.xlsx.
Uruchamiany codziennie przez GitHub Actions.
"""

import os, sys, json, base64, urllib.request, urllib.parse
from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Konfiguracja z GitHub Secrets ───────────────────────────────────────────
CLIENT_ID     = os.environ.get("ALLEGRO_CLIENT_ID", "")
CLIENT_SECRET = os.environ.get("ALLEGRO_CLIENT_SECRET", "")
EXCEL_FILE    = "LEGO_Monitor.xlsx"

if not CLIENT_ID or not CLIENT_SECRET:
    print("BLAD: Brak ALLEGRO_CLIENT_ID lub ALLEGRO_CLIENT_SECRET.")
    sys.exit(1)

# ── Style ────────────────────────────────────────────────────────────────────
C_LIGHT  = "D6EAF8"
C_WHITE  = "FFFFFF"
C_BORDER = "BDC3C7"

def _border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _style_row(ws, row, ncols, alt):
    bg = C_LIGHT if alt else C_WHITE
    for col in range(1, ncols + 1):
        c = ws.cell(row, col)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.font      = Font(name="Arial", size=10)
        c.border    = _border()
        c.alignment = Alignment(horizontal="center", vertical="center")

# ── Allegro: token ───────────────────────────────────────────────────────────
def get_token() -> str:
    creds = base64.b64encode(f"{CLIENT_ID}:{CLIENT_SECRET}".encode()).decode()
    data  = urllib.parse.urlencode({"grant_type": "client_credentials"}).encode()
    req   = urllib.request.Request(
        "https://allegro.pl/auth/oauth/token", data=data,
        headers={"Authorization": f"Basic {creds}",
                 "Content-Type": "application/x-www-form-urlencoded"})
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read())["access_token"]

# ── Allegro: oferty dla konkretnego numeru setu ──────────────────────────────
def fetch_offers(token: str, set_nr: str) -> list:
    params = urllib.parse.urlencode({
        "phrase":   f"LEGO {set_nr}",
        "limit":    100,
        "sort":     "price",
        "include":  "-all",
        "fallback": "false",
    })
    req = urllib.request.Request(
        f"https://api.allegro.pl/offers/listing?{params}",
        headers={"Authorization": f"Bearer {token}",
                 "Accept": "application/vnd.allegro.public.v1+json"})
    with urllib.request.urlopen(req) as r:
        data = json.loads(r.read())
    return (data.get("items", {}).get("regular", []) +
            data.get("items", {}).get("promoted", []))

# ── Parsowanie danych z ofert ────────────────────────────────────────────────
def parse_stats(items: list) -> dict:
    prices      = []
    bought_last = 0

    for it in items:
        # Cena
        try:
            prices.append(float(it["sellingMode"]["price"]["amount"]))
        except (KeyError, TypeError, ValueError):
            pass

        # "X osób kupiło ostatnio" – pole popularityScore lub boughtCount
        stats = it.get("stats", {})
        bought_last = max(
            bought_last,
            stats.get("boughtCount", 0) or 0,
            stats.get("popularityScore", 0) or 0,
        )

    return {
        "min_price":   round(min(prices), 2) if prices else None,
        "offer_count": len(prices),
        "bought_last": bought_last,
    }

# ── Odczyt listy zestawów z arkusza "Lista zestawow" ────────────────────────
def read_sets(wb) -> list:
    ws   = wb["Lista zestawow"]
    sets = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        nr    = str(row[0]).strip() if row[0] else ""
        name  = str(row[1]).strip() if row[1] else ""
        seria = str(row[2]).strip() if row[2] else ""
        aktyw = str(row[4]).strip().upper() if len(row) > 4 and row[4] else ""
        if nr and aktyw == "TAK":
            sets.append((nr, name, seria))
    return sets

# ── Zapis wiersza do "Dane dzienne" ─────────────────────────────────────────
def append_row(wb, today, nr, name, seria, stats):
    ws       = wb["Dane dzienne"]
    next_row = ws.max_row + 1
    alt      = next_row % 2 == 0

    ws.cell(next_row, 1, today)
    ws.cell(next_row, 2, nr)
    ws.cell(next_row, 3, name)
    ws.cell(next_row, 4, seria)
    ws.cell(next_row, 5, stats["min_price"])
    ws.cell(next_row, 6, stats["offer_count"])
    ws.cell(next_row, 7, stats["bought_last"])
    ws.cell(next_row, 8, "OK")

    _style_row(ws, next_row, 8, alt)
    if stats["min_price"]:
        ws.cell(next_row, 5).number_format = '#,##0.00 "zł"'

# ── Aktualizacja tabeli pomocniczej w arkuszu wykresu ───────────────────────
def update_chart_table(wb, today, nr, name, seria, stats):
    ws = wb["Wykres trendu"]

    # Znajdź kolumnę dla tego setu (nagłówki w wierszu 3, kolumny 2-5)
    col_map = {}
    for col in range(2, 10):
        val = ws.cell(3, col).value
        if val and str(val).startswith(nr):
            col_map[nr] = col
            break

    if nr not in col_map:
        return  # set nie ma jeszcze kolumny w wykresie – pomijamy

    # Sprawdź czy dzisiejszy wiersz już istnieje
    target_row = None
    for row in range(4, ws.max_row + 2):
        if ws.cell(row, 1).value == today:
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1
        ws.cell(target_row, 1, today)

    ws.cell(target_row, col_map[nr], stats["min_price"])
    if stats["min_price"]:
        ws.cell(target_row, col_map[nr]).number_format = '#,##0.00 "zł"'

# ── MAIN ─────────────────────────────────────────────────────────────────────
def main():
    today      = str(date.today())
    excel_path = Path(EXCEL_FILE)

    if not excel_path.exists():
        print(f"BLAD: Nie znaleziono {EXCEL_FILE}")
        sys.exit(1)

    print(f"[{today}] Pobieram token Allegro...")
    token = get_token()
    print("Token OK.")

    wb   = openpyxl.load_workbook(excel_path)
    sets = read_sets(wb)

    if not sets:
        print("Brak aktywnych zestawow w arkuszu 'Lista zestawow'.")
        sys.exit(0)

    print(f"Znaleziono {len(sets)} aktywnych zestawow do sledzenia.")

    for nr, name, seria in sets:
        print(f"  -> {nr} {name}...", end=" ", flush=True)
        try:
            items = fetch_offers(token, nr)
            stats = parse_stats(items)
            append_row(wb, today, nr, name, seria, stats)
            update_chart_table(wb, today, nr, name, seria, stats)
            print(f"cena min: {stats['min_price']} zl | ofert: {stats['offer_count']} | kupilo ostatnio: {stats['bought_last']}")
        except Exception as e:
            print(f"BLAD: {e}")
            append_row(wb, today, nr, name, seria,
                       {"min_price": None, "offer_count": 0, "bought_last": 0})

    wb.save(excel_path)
    print(f"\nZapisano: {excel_path}")

if __name__ == "__main__":
    main()
